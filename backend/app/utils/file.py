import asyncio
import os
import tempfile
from io import BytesIO

from configs.load_env import PROJECT_ROOT, db_path
from fastapi import UploadFile
from models.user import Feedback
from utils import db


def safe_filename(filename: str) -> str:
    """
    去除路径分隔符，只保留文件名本身，防止路径穿越（如 ../../etc/passwd）
    """
    name = os.path.basename(filename.replace('\\', '/'))
    if name in ('', '.', '..'):
        raise ValueError(f'invalid filename: {filename!r}')
    return name


def get_folders_list(root_dir: str) -> list:
    folders_list = []
    dir = os.path.join(PROJECT_ROOT, root_dir)
    for dirpath, dirnames, filenames in os.walk(dir):
        for dirname in dirnames:
            folders_list.append(dirname)
    return folders_list


async def save_feedback(client_ip: str, feedback: Feedback):
    await asyncio.to_thread(db.save_feedback, db_path, client_ip, feedback.email, feedback.message)


def _read_file_sync(file: UploadFile) -> str:
    """同步读取文件内容（用于 run_in_executor 包装）"""
    filename = file.filename or ''
    ext = filename.split('.')[-1].lower() if '.' in filename else ''

    if ext == 'docx':
        from docx import Document
        with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as temp_file:
            temp_path = temp_file.name
            temp_file.write(file.file.read())
        try:
            doc = Document(temp_path)
            content_parts = []
            for paragraph in doc.paragraphs:
                content_parts.append(paragraph.text)
            content = ' '.join(content_parts)
        finally:
            os.unlink(temp_path)

    elif ext == 'pdf':
        import pdfplumber
        with pdfplumber.open(BytesIO(file.file.read())) as pdf:
            content = ''
            for page in pdf.pages:
                page_text = page.extract_text() or ''
                content = content + page_text

    elif ext == 'xlsx':
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file.file.read()), read_only=True, data_only=True)
        content_parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_text = ' '.join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    content_parts.append(row_text)
        wb.close()
        content = ' '.join(content_parts)

    else:
        contents = file.file.read()
        try:
            content = contents.decode('utf-8')
        except UnicodeDecodeError:
            content = contents.decode('gbk')

    return ' '.join(content.split())


async def read_file_contents(file: UploadFile) -> str:
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _read_file_sync, file)
