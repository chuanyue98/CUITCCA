""".xlsx：openpyxl 逐 sheet 逐行提取，保留行结构。

统一给生产摄取链路和 QA 生成上传链路用（原来两边分别通过
``SimpleDirectoryReader``/``PandasExcelReader`` 和直接调 openpyxl 各实现一遍，
这里合并成一个实现，消除重复）。每个 sheet 产出一段独立文本（带
``# sheet 名`` 小标题 + 逐行 ``" | "`` 连接的单元格），而不是把所有 sheet
拍平成一整块——这样 ``documents_from_file`` 能继续保留"一个 sheet 一个
Document"的粒度（和以前 ``PandasExcelReader(concat_rows=True)`` 多 sheet 时
的行为一致），检索时不会出现"一个 chunk 横跨两个毫不相关的 sheet"的问题。

``read_only=True, data_only=True``：只读公式计算后的值，不加载单元格样式，
大表格也能快速读完，避免内存/耗时问题。
"""
from __future__ import annotations

import openpyxl

from ._source import Source, as_path_or_stream
from .types import ParseResult


def parse_xlsx(source: Source) -> ParseResult:
    try:
        wb = openpyxl.load_workbook(as_path_or_stream(source), read_only=True, data_only=True)
    except Exception as e:
        return ParseResult.failure(f"{type(e).__name__}: {e}")

    texts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines = [f"# {sheet_name}"]
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) for cell in row if cell is not None]
            if cells:
                lines.append(" | ".join(cells))
        if len(lines) > 1:  # 有实际数据行，而不是只有个标题
            texts.append("\n".join(lines))
    wb.close()

    return ParseResult.success(texts)
